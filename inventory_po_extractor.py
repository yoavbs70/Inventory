import re
import sys
from datetime import date, datetime
from pathlib import Path

VERSION = "1.0.7"

import openpyxl
from openpyxl import Workbook

SOURCE = Path(r"C:\Claude Projects\Inventory\InventoryTransactionsF6.xlsx")
SYNC_SCRIPT = SOURCE.parent / "sync-inventory.ps1"
TASK_NAME = "SyncInventoryFromSharePoint"
PO_PATTERN = re.compile(r'[PN][A-Z]*[-.](\d{5,})')


def parse_date(s):
    for fmt in ('%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date format: '{s}'. Use dd/mm/yyyy.")


def to_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        try:
            return parse_date(val)
        except ValueError:
            return None
    return None


def prompt(msg, default=None):
    val = input(msg).strip()
    return val if val else default


def setup_scheduled_task():
    time_str = prompt("Run time [09:00]: ", default="09:00")
    if not re.fullmatch(r'\d{2}:\d{2}', time_str):
        print("Invalid time format. Use HH:MM.")
        sys.exit(1)

    ps_content = f"""\
$action  = New-ScheduledTaskAction -Execute "powershell.exe" `
           -Argument '-NonInteractive -WindowStyle Hidden -ExecutionPolicy Bypass -File "{SYNC_SCRIPT}"'
$trigger = New-ScheduledTaskTrigger -Daily -At "{time_str}"
Register-ScheduledTask -TaskName "{TASK_NAME}" -Action $action -Trigger $trigger -Force
Write-Host "Task '{TASK_NAME}' scheduled daily at {time_str}."
Read-Host "Press Enter to close"
"""

    out_path = SOURCE.parent / "setup-task.ps1"
    out_path.write_text(ps_content, encoding="utf-8")

    print(f"\nScript written to:\n  {out_path}")
    print("\nRun it as Administrator to create/update the scheduled task:")
    print(f"  Right-click -> Run with PowerShell")


def main():
    print(f"Inventory PO Extractor v{VERSION}")
    print()
    print("1. Search Catalog")
    print("2. Setup scheduled sync (Task Scheduler)")
    choice = prompt("\nChoice: ")
    if choice == "2":
        setup_scheduled_task()
        return
    if choice != "1":
        print("Invalid choice.")
        sys.exit(1)

    print()
    catalog = prompt("Catalog: ")
    if not catalog:
        print("Catalog is required.")
        sys.exit(1)

    from_str = prompt("From date (dd/mm/yyyy): ")
    if not from_str:
        print("From date is required.")
        sys.exit(1)
    try:
        from_date = parse_date(from_str)
    except ValueError as e:
        print(e)
        sys.exit(1)

    to_str = prompt("To date   (dd/mm/yyyy) [Enter = today]: ")
    if to_str:
        try:
            to_date_val = parse_date(to_str)
        except ValueError as e:
            print(e)
            sys.exit(1)
    else:
        to_date_val = date.today()

    print(f"\nSearching: Catalog={catalog}  |  {from_date} → {to_date_val}\n")

    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        catalog_val = row[2]   # column C
        notes_val   = row[6]   # column G
        date_val    = row[8]   # column I

        if not catalog_val or str(catalog_val).strip().lower() != catalog.lower():
            continue

        txn_date = to_date(date_val)
        if txn_date is None or not (from_date <= txn_date <= to_date_val):
            continue

        if not notes_val:
            continue
        match = PO_PATTERN.search(str(notes_val))
        if not match:
            continue

        results.append((txn_date, list(row), match.group(1)))

    if not results:
        print("No matching records found.")
        sys.exit(0)

    results.sort(key=lambda x: (x[0], int(x[2])))

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "PO Results"
    out_ws.append(headers + ["PO Number"])
    for _, row, po in results:
        out_ws.append(row + [po])

    out_name = f"PO_Results_{catalog}_{from_date.strftime('%Y%m%d')}_{to_date_val.strftime('%Y%m%d')}.xlsx"
    out_path = SOURCE.parent / out_name
    out_wb.save(out_path)

    print(f"{len(results)} record(s) saved to:\n{out_path}")


if __name__ == "__main__":
    main()
